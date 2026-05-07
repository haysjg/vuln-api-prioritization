#!/usr/bin/env python3
"""
CVE Prioritization Script for CrowdStrike Falcon

This script retrieves vulnerabilities from CrowdStrike Falcon and calculates
a priority score based on multiple factors:
- CVSS severity score
- CrowdStrike Expert Rating (CTI-based threat intelligence)
- Exploitation status (in-the-wild, PoC available, etc.)
- Number of affected hosts
- Asset criticality (based on host groups and tags)
- Age of the vulnerability

Output: Prioritized list of CVEs exported to CSV, JSON, and Excel formats.
"""

import os
import sys
import json
import argparse
from datetime import datetime, timezone
from typing import Dict, List, Any
from collections import defaultdict

from falconpy import Intel, Hosts

# Configurable weights for prioritization scoring
SCORING_WEIGHTS = {
    'cvss_weight': 10,              # Multiplier for CVSS base score (0-10) -> 0-100
    'exploit_status_weight': 30,     # Points per exploit status level
    'hosts_affected_weight': 0.5,    # Points per affected host
    'asset_criticality_weight': 50,  # Bonus for critical assets
    'age_weight': 0.1,               # Points per day since publication
    'expert_rating_weight': 80       # CrowdStrike Expert Rating (CTI analysis)
}

# Expert Rating scoring (CrowdStrike Threat Intelligence)
EXPERT_RATING_SCORES = {
    'CRITICAL': 100,
    'HIGH': 70,
    'MEDIUM': 40,
    'LOW': 10,
    'INFORMATIONAL': 0,
    'N/A': 0,
    '': 0,
    None: 0
}

# Asset criticality levels based on tags and groups
CRITICAL_ASSET_INDICATORS = {
    'groups': ['production', 'prod', 'critical', 'dmz', 'domain-controllers', 'pci', 'financial'],
    'tags': ['crown-jewels', 'tier-0', 'pci-dss', 'sox', 'hipaa', 'critical', 'production']
}


def parse_arguments():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(
        description='Prioritize CrowdStrike Falcon vulnerabilities',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )

    parser.add_argument(
        '--filter',
        type=str,
        help='FQL filter for vulnerability query (e.g., "cve.severity:\'CRITICAL\'+cve.severity:\'HIGH\'")',
        default=None
    )

    parser.add_argument(
        '--min-score',
        type=float,
        help='Minimum CVSS base score to include (default: 0.0)',
        default=0.0
    )

    parser.add_argument(
        '--limit',
        type=int,
        help='Maximum number of vulnerabilities to process (default: all)',
        default=None
    )

    parser.add_argument(
        '--output-dir',
        type=str,
        help='Output directory for reports (default: current directory)',
        default='.'
    )

    parser.add_argument(
        '--output-prefix',
        type=str,
        help='Prefix for output files (default: vulnerability_prioritization)',
        default='vulnerability_prioritization'
    )

    parser.add_argument(
        '--top',
        type=int,
        help='Show only top N vulnerabilities (default: all)',
        default=None
    )

    parser.add_argument(
        '--verbose',
        action='store_true',
        help='Enable verbose output'
    )

    return parser.parse_args()


def calculate_exploit_status_score(exploit_status: int) -> int:
    """
    Calculate score based on exploitation status

    Exploit Status Values:
    0 = Not exploited
    1 = PoC exists
    2 = Public exploit available
    3 = Active exploitation observed
    4 = Widespread exploitation (e.g., ransomware)
    """
    exploit_scores = {
        0: 0,
        1: 30,
        2: 60,
        3: 90,
        4: 120
    }
    return exploit_scores.get(exploit_status, 0)


def calculate_asset_criticality(host_info: Dict) -> tuple:
    """
    Determine if an asset is critical based on groups and tags
    Returns (is_critical: bool, criticality_score: int, reasons: list)
    """
    is_critical = False
    reasons = []
    score = 0

    # Check host groups
    groups = host_info.get('groups', [])
    for group in groups:
        group_lower = group.lower()
        for critical_indicator in CRITICAL_ASSET_INDICATORS['groups']:
            if critical_indicator in group_lower:
                is_critical = True
                score += 50
                reasons.append(f"Critical group: {group}")
                break

    # Check host tags
    tags = host_info.get('tags', [])
    for tag in tags:
        tag_lower = tag.lower()
        for critical_indicator in CRITICAL_ASSET_INDICATORS['tags']:
            if critical_indicator in tag_lower:
                is_critical = True
                score += 50
                reasons.append(f"Critical tag: {tag}")
                break

    return is_critical, min(score, 100), reasons  # Cap at 100


def calculate_age_days(published_date: str) -> int:
    """Calculate number of days since CVE was published"""
    if not published_date:
        return 0

    try:
        # Handle different datetime formats
        if 'T' in published_date:
            pub_dt = datetime.fromisoformat(published_date.replace('Z', '+00:00'))
        else:
            pub_dt = datetime.strptime(published_date, '%Y-%m-%d')
            pub_dt = pub_dt.replace(tzinfo=timezone.utc)

        now = datetime.now(timezone.utc)
        delta = now - pub_dt
        return delta.days
    except Exception as e:
        return 0


def calculate_priority_score(vuln: Dict, host_details: Dict = None) -> Dict:
    """
    Calculate comprehensive priority score for a vulnerability

    Returns dict with:
    - total_score: Overall priority score
    - component_scores: Breakdown of score components
    - metadata: Additional context for prioritization
    """
    cve = vuln.get('cve', {})
    apps = vuln.get('apps', [])
    host_info = vuln.get('host_info', {})

    # Component scores
    cvss_score = float(cve.get('base_score', 0))
    cvss_points = cvss_score * SCORING_WEIGHTS['cvss_weight']

    exploit_status = int(cve.get('exploit_status', 0))
    exploit_points = calculate_exploit_status_score(exploit_status) * (SCORING_WEIGHTS['exploit_status_weight'] / 30)

    # Number of hosts affected
    hosts_affected = 0
    for app in apps:
        hosts_affected += int(app.get('number_of_hosts_affected', 0))
    if hosts_affected == 0 and vuln.get('aid'):
        hosts_affected = 1
    hosts_points = hosts_affected * SCORING_WEIGHTS['hosts_affected_weight']

    # Asset criticality
    is_critical, criticality_score, criticality_reasons = calculate_asset_criticality(host_info)
    asset_points = criticality_score * (SCORING_WEIGHTS['asset_criticality_weight'] / 100)

    # Age factor
    published_date = cve.get('published_date') or cve.get('created_timestamp')
    age_days = calculate_age_days(published_date)
    age_points = age_days * SCORING_WEIGHTS['age_weight']

    # CrowdStrike Expert Rating (CTI-based)
    expert_rating = cve.get('exprt_rating', 'N/A')
    expert_rating_value = calculate_expert_rating_score(expert_rating)
    expert_points = expert_rating_value * (SCORING_WEIGHTS['expert_rating_weight'] / 100)

    # Calculate total score
    total_score = cvss_points + exploit_points + hosts_points + asset_points + age_points + expert_points

    return {
        'total_score': round(total_score, 2),
        'component_scores': {
            'cvss': round(cvss_points, 2),
            'expert_rating': round(expert_points, 2),
            'exploitation': round(exploit_points, 2),
            'exposure': round(hosts_points, 2),
            'asset_criticality': round(asset_points, 2),
            'age': round(age_points, 2)
        },
        'metadata': {
            'cvss_base_score': cvss_score,
            'severity': cve.get('severity', 'UNKNOWN'),
            'expert_rating': expert_rating,
            'expert_rating_score': expert_rating_value,
            'exploit_status': exploit_status,
            'exploit_status_label': get_exploit_status_label(exploit_status),
            'hosts_affected': hosts_affected,
            'is_critical_asset': is_critical,
            'criticality_reasons': criticality_reasons,
            'age_days': age_days,
            'published_date': published_date,
            'cve_id': cve.get('id', 'Unknown'),
            'description': cve.get('description', '')[:200],  # Truncate
            'vendor_advisory': cve.get('vendor_advisory', []),
            'exprt_rating': expert_rating  # Keep for backwards compatibility
        }
    }


def get_exploit_status_label(exploit_status: int) -> str:
    """Get human-readable label for exploit status"""
    labels = {
        0: 'Not Exploited',
        1: 'PoC Exists',
        2: 'Public Exploit Available',
        3: 'Active Exploitation',
        4: 'Widespread Exploitation'
    }
    return labels.get(exploit_status, 'Unknown')


def calculate_expert_rating_score(expert_rating: str) -> int:
    """
    Calculate score based on CrowdStrike Expert Rating

    Expert Rating is CrowdStrike's CTI-based assessment that considers:
    - Active exploitation in the wild
    - Threat actor usage
    - Ease of exploitation
    - Impact potential
    - Available mitigations

    Returns score 0-100 based on rating level
    """
    if not expert_rating:
        return 0

    # Normalize rating (handle case variations)
    rating_upper = str(expert_rating).upper().strip()

    # Direct mapping
    if rating_upper in EXPERT_RATING_SCORES:
        return EXPERT_RATING_SCORES[rating_upper]

    # Handle partial matches or variations
    if 'CRITICAL' in rating_upper:
        return EXPERT_RATING_SCORES['CRITICAL']
    elif 'HIGH' in rating_upper:
        return EXPERT_RATING_SCORES['HIGH']
    elif 'MEDIUM' in rating_upper or 'MODERATE' in rating_upper:
        return EXPERT_RATING_SCORES['MEDIUM']
    elif 'LOW' in rating_upper:
        return EXPERT_RATING_SCORES['LOW']
    elif 'INFO' in rating_upper:
        return EXPERT_RATING_SCORES['INFORMATIONAL']

    # Default to 0 for unknown ratings
    return 0



def query_all_vulnerabilities(intel_api: Intel, fql_filter: str = None, limit: int = None, verbose: bool = False) -> List[str]:
    """Query all vulnerability IDs with pagination"""
    all_vuln_ids = []
    offset = None
    page = 1

    if verbose:
        print(f"Querying vulnerability IDs (filter: {fql_filter or 'none'})...")

    while True:
        if verbose:
            print(f"  Fetching page {page}...", end='', flush=True)

        response = intel_api.QueryVulnerabilities(
            filter=fql_filter,
            offset=offset,
            limit=5000
        )

        if response['status_code'] != 200:
            print(f"\nError querying vulnerabilities: {response['body'].get('errors', [])}")
            break

        vuln_ids = response['body']['resources']
        all_vuln_ids.extend(vuln_ids)

        if verbose:
            print(f" {len(vuln_ids)} IDs retrieved (total: {len(all_vuln_ids)})")

        # Check if we've reached the limit
        if limit and len(all_vuln_ids) >= limit:
            all_vuln_ids = all_vuln_ids[:limit]
            if verbose:
                print(f"  Reached limit of {limit} vulnerabilities")
            break

        # Check for more pages
        offset = response['body']['meta']['pagination'].get('offset')
        if not offset:
            break

        page += 1

    if verbose:
        print(f"Total vulnerability IDs retrieved: {len(all_vuln_ids)}\n")

    return all_vuln_ids


def get_vulnerability_details(intel_api: Intel, vuln_ids: List[str], verbose: bool = False) -> List[Dict]:
    """Retrieve detailed information for vulnerabilities in batches"""
    all_vulns = []
    batch_size = 200  # API limit per request

    if verbose:
        print(f"Retrieving detailed vulnerability information...")

    for i in range(0, len(vuln_ids), batch_size):
        batch = vuln_ids[i:i+batch_size]

        if verbose:
            print(f"  Processing batch {i//batch_size + 1}/{(len(vuln_ids)-1)//batch_size + 1} ({len(batch)} items)...", end='', flush=True)

        response = intel_api.GetVulnerabilities(ids=batch)

        if response['status_code'] != 200:
            print(f"\nError getting vulnerability details: {response['body'].get('errors', [])}")
            continue

        vulns = response['body']['resources']
        all_vulns.extend(vulns)

        if verbose:
            print(f" Done (total: {len(all_vulns)})")

    if verbose:
        print(f"Total vulnerabilities retrieved: {len(all_vulns)}\n")

    return all_vulns


def prioritize_vulnerabilities(vulns: List[Dict], min_score: float = 0.0, verbose: bool = False) -> List[Dict]:
    """Calculate priority scores for all vulnerabilities and sort by priority"""
    if verbose:
        print("Calculating priority scores...")

    prioritized = []

    for vuln in vulns:
        cve = vuln.get('cve', {})
        cvss_score = float(cve.get('base_score', 0))

        # Skip if below minimum CVSS score
        if cvss_score < min_score:
            continue

        priority_info = calculate_priority_score(vuln)

        prioritized.append({
            'vulnerability': vuln,
            'priority': priority_info
        })

    # Sort by total priority score (descending)
    prioritized.sort(key=lambda x: x['priority']['total_score'], reverse=True)

    if verbose:
        print(f"Prioritized {len(prioritized)} vulnerabilities\n")

    return prioritized


def export_to_csv(prioritized_vulns: List[Dict], output_file: str):
    """Export prioritized vulnerabilities to CSV"""
    import csv

    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Header
        writer.writerow([
            'Rank',
            'Priority Score',
            'CVE ID',
            'CVSS Score',
            'Severity',
            'Expert Rating',
            'Exploit Status',
            'Hosts Affected',
            'Critical Asset',
            'Age (days)',
            'Published Date',
            'Description',
            'CVSS Points',
            'Expert Rating Points',
            'Exploit Points',
            'Exposure Points',
            'Asset Points',
            'Age Points'
        ])

        # Data rows
        for rank, item in enumerate(prioritized_vulns, 1):
            priority = item['priority']
            meta = priority['metadata']
            scores = priority['component_scores']

            writer.writerow([
                rank,
                priority['total_score'],
                meta['cve_id'],
                meta['cvss_base_score'],
                meta['severity'],
                meta['expert_rating'],
                meta['exploit_status_label'],
                meta['hosts_affected'],
                'YES' if meta['is_critical_asset'] else 'NO',
                meta['age_days'],
                meta['published_date'],
                meta['description'],
                scores['cvss'],
                scores['expert_rating'],
                scores['exploitation'],
                scores['exposure'],
                scores['asset_criticality'],
                scores['age']
            ])

    print(f"✓ CSV report saved: {output_file}")


def export_to_json(prioritized_vulns: List[Dict], output_file: str):
    """Export prioritized vulnerabilities to JSON"""
    output_data = {
        'generated_at': datetime.now(timezone.utc).isoformat(),
        'total_count': len(prioritized_vulns),
        'scoring_weights': SCORING_WEIGHTS,
        'vulnerabilities': []
    }

    for rank, item in enumerate(prioritized_vulns, 1):
        output_data['vulnerabilities'].append({
            'rank': rank,
            'priority_score': item['priority']['total_score'],
            'score_breakdown': item['priority']['component_scores'],
            'metadata': item['priority']['metadata'],
            'raw_data': item['vulnerability']
        })

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, indent=2, default=str)

    print(f"✓ JSON report saved: {output_file}")


def export_to_excel(prioritized_vulns: List[Dict], output_file: str):
    """Export prioritized vulnerabilities to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Prioritized CVEs"

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Headers
        headers = [
            'Rank', 'Priority Score', 'CVE ID', 'CVSS', 'Severity', 'Expert Rating',
            'Exploit Status', 'Hosts', 'Critical Asset', 'Age (days)',
            'Published', 'Description',
            'CVSS Pts', 'Expert Pts', 'Exploit Pts', 'Exposure Pts', 'Asset Pts', 'Age Pts'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        for rank, item in enumerate(prioritized_vulns, 1):
            priority = item['priority']
            meta = priority['metadata']
            scores = priority['component_scores']

            row = rank + 1
            ws.cell(row=row, column=1, value=rank)
            ws.cell(row=row, column=2, value=priority['total_score'])
            ws.cell(row=row, column=3, value=meta['cve_id'])
            ws.cell(row=row, column=4, value=meta['cvss_base_score'])
            ws.cell(row=row, column=5, value=meta['severity'])
            ws.cell(row=row, column=6, value=meta['expert_rating'])
            ws.cell(row=row, column=7, value=meta['exploit_status_label'])
            ws.cell(row=row, column=8, value=meta['hosts_affected'])
            ws.cell(row=row, column=9, value='YES' if meta['is_critical_asset'] else 'NO')
            ws.cell(row=row, column=10, value=meta['age_days'])
            ws.cell(row=row, column=11, value=meta['published_date'])
            ws.cell(row=row, column=12, value=meta['description'])
            ws.cell(row=row, column=13, value=scores['cvss'])
            ws.cell(row=row, column=14, value=scores['expert_rating'])
            ws.cell(row=row, column=15, value=scores['exploitation'])
            ws.cell(row=row, column=16, value=scores['exposure'])
            ws.cell(row=row, column=17, value=scores['asset_criticality'])
            ws.cell(row=row, column=18, value=scores['age'])

            # Color-code by severity
            severity_colors = {
                'CRITICAL': 'FFC7CE',
                'HIGH': 'FFD9B3',
                'MEDIUM': 'FFEB9C',
                'LOW': 'C6EFCE'
            }
            if meta['severity'] in severity_colors:
                fill = PatternFill(start_color=severity_colors[meta['severity']],
                                 end_color=severity_colors[meta['severity']],
                                 fill_type="solid")
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = fill

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(output_file)
        print(f"✓ Excel report saved: {output_file}")

    except ImportError:
        print("⚠ Warning: openpyxl not installed. Skipping Excel export.")
        print("  Install with: pip install openpyxl")


def print_summary(prioritized_vulns: List[Dict], top_n: int = None):
    """Print summary of top vulnerabilities to console"""
    print("=" * 100)
    print("VULNERABILITY PRIORITIZATION SUMMARY")
    print("=" * 100)

    if top_n:
        display_vulns = prioritized_vulns[:top_n]
        print(f"\nShowing top {top_n} of {len(prioritized_vulns)} total vulnerabilities\n")
    else:
        display_vulns = prioritized_vulns
        print(f"\nTotal vulnerabilities: {len(prioritized_vulns)}\n")

    # Summary statistics
    severity_counts = defaultdict(int)
    exploit_counts = defaultdict(int)
    expert_rating_counts = defaultdict(int)
    critical_asset_count = 0
    total_hosts_affected = 0

    for item in prioritized_vulns:
        meta = item['priority']['metadata']
        severity_counts[meta['severity']] += 1
        exploit_counts[meta['exploit_status_label']] += 1
        expert_rating_counts[meta.get('expert_rating', 'N/A')] += 1
        if meta['is_critical_asset']:
            critical_asset_count += 1
        total_hosts_affected += meta['hosts_affected']

    print(f"Severity Distribution:")
    for severity in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW', 'UNKNOWN']:
        count = severity_counts.get(severity, 0)
        if count > 0:
            print(f"  {severity}: {count}")

    print(f"\nExpert Rating Distribution:")
    for rating in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW', 'INFORMATIONAL', 'N/A']:
        count = expert_rating_counts.get(rating, 0)
        if count > 0:
            print(f"  {rating}: {count}")

    print(f"\nExploitation Status:")
    for status, count in sorted(exploit_counts.items(), reverse=True):
        print(f"  {status}: {count}")

    print(f"\nCritical Assets Affected: {critical_asset_count}")
    print(f"Total Host Instances Affected: {total_hosts_affected}")

    # Top vulnerabilities
    print("\n" + "=" * 120)
    print("TOP PRIORITY VULNERABILITIES")
    print("=" * 120)
    print(f"\n{'#':<4} {'Score':<8} {'CVE ID':<18} {'CVSS':<6} {'Severity':<10} {'Expert':<10} {'Exploit Status':<25} {'Hosts':<6} {'Critical':<9}")
    print("-" * 120)

    for rank, item in enumerate(display_vulns, 1):
        priority = item['priority']
        meta = priority['metadata']

        print(f"{rank:<4} {priority['total_score']:<8.2f} {meta['cve_id']:<18} "
              f"{meta['cvss_base_score']:<6.1f} {meta['severity']:<10} "
              f"{meta.get('expert_rating', 'N/A'):<10} "
              f"{meta['exploit_status_label']:<25} {meta['hosts_affected']:<6} "
              f"{'YES' if meta['is_critical_asset'] else 'NO':<9}")

    print("\n" + "=" * 120)


def main():
    args = parse_arguments()

    # Authentication
    client_id = os.getenv('FALCON_CLIENT_ID')
    client_secret = os.getenv('FALCON_CLIENT_SECRET')

    if not client_id or not client_secret:
        print("Error: FALCON_CLIENT_ID and FALCON_CLIENT_SECRET environment variables must be set")
        sys.exit(1)

    print("CrowdStrike Falcon - CVE Prioritization Tool")
    print("=" * 100)
    print()

    # Initialize API clients
    if args.verbose:
        print("Authenticating with CrowdStrike Falcon...")

    intel_api = Intel(client_id=client_id, client_secret=client_secret)

    # Query vulnerabilities
    vuln_ids = query_all_vulnerabilities(
        intel_api,
        fql_filter=args.filter,
        limit=args.limit,
        verbose=args.verbose
    )

    if not vuln_ids:
        print("No vulnerabilities found matching the criteria.")
        sys.exit(0)

    # Get detailed information
    vulnerabilities = get_vulnerability_details(intel_api, vuln_ids, verbose=args.verbose)

    # Calculate priorities
    prioritized = prioritize_vulnerabilities(
        vulnerabilities,
        min_score=args.min_score,
        verbose=args.verbose
    )

    if not prioritized:
        print(f"No vulnerabilities found with CVSS score >= {args.min_score}")
        sys.exit(0)

    # Generate timestamp for filenames
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # Export to multiple formats
    output_base = os.path.join(args.output_dir, f"{args.output_prefix}_{timestamp}")

    export_to_csv(prioritized, f"{output_base}.csv")
    export_to_json(prioritized, f"{output_base}.json")
    export_to_excel(prioritized, f"{output_base}.xlsx")

    # Print summary
    print()
    print_summary(prioritized, top_n=args.top)

    print("\n✓ Prioritization complete!")


if __name__ == "__main__":
    main()
